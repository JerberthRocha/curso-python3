"""
https://openpyxl.readthedocs.io/en/stable/
"""
import openpyxl
from random import uniform

planilha = openpyxl.Workbook()
planilha.create_sheet('planilha1', 0)
planilha.create_sheet('planilha2', 1)

planilha1 = planilha['planilha1']
planilha2 = planilha['planilha2']

for linha in range(1, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(1, 16):
    planilha2.cell(linha, 1).value = f'Jerberth {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 2).value = f'Rocha {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 3).value = f'Silva {linha} {round(uniform(10, 100), 2)}'

planilha.save('nova_planilha.xlsx')

"""pedidos = openpyxl.load_workbook('planilha.xlsx')
nome_planilhas = pedidos.sheetnames
planilha1 = pedidos['Página1']

for linha in range(5, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

# planilha1['B3'].value = 20220
pedidos.save('nova_planilha.xlsx')"""

# print(planilha1['b6'].value)
# for campo in planilha1['b']:

# for linha in planilha1['a1:c2']:
#     for coluna in linha:
#         print(coluna.value)

# for linha in planilha1:
#     if linha[0].value is not None:
#         print(linha[0].value, end=" ")
#     if linha[1].value is not None:
#         print(linha[1].value, end=" ")
#     if linha[2].value is not None:
#         print(linha[2].value)

    